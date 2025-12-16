"""
Export Service - 报表导出服务
支持 Excel、PDF、CSV 格式导出
"""
import os
import csv
from datetime import datetime
from io import BytesIO

# Excel 导出
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF 导出
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ExportService:
    """报表导出服务"""

    def __init__(self, output_dir=None):
        """初始化导出服务"""
        if output_dir:
            self.output_dir = output_dir
        else:
            # 默认输出目录
            self.output_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                'reports'
            )

        # 确保目录存在
        os.makedirs(self.output_dir, exist_ok=True)

    def export_to_excel(self, report_data, report_type, report_name=None):
        """
        导出为 Excel 文件

        Args:
            report_data: 报表数据字典
            report_type: 报表类型
            report_name: 报表名称（可选）

        Returns:
            str: 生成的文件路径
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl 未安装，无法生成 Excel 文件")

        wb = Workbook()
        ws = wb.active

        # 设置标题
        title = report_name or self._get_report_title(report_type)
        ws.title = title[:31]  # Excel sheet name 最长 31 字符

        # 样式定义
        title_font = Font(name='Microsoft YaHei', size=16, bold=True)
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell_font = Font(name='Microsoft YaHei', size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        row = 1

        # 写入标题
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=title).font = title_font
        ws.cell(row=row, column=1).alignment = center_align
        row += 2

        # 写入日期范围
        if 'date_range' in report_data:
            dr = report_data['date_range']
            ws.cell(row=row, column=1, value=f"日期范围: {dr.get('start', '')} 至 {dr.get('end', '')}")
            row += 2

        # 根据报表类型生成不同的内容
        if report_type == 'production':
            row = self._write_production_excel(ws, report_data, row, header_font, header_fill, cell_font, thin_border, center_align, left_align)
        elif report_type == 'order':
            row = self._write_order_excel(ws, report_data, row, header_font, header_fill, cell_font, thin_border, center_align, left_align)
        elif report_type == 'task':
            row = self._write_task_excel(ws, report_data, row, header_font, header_fill, cell_font, thin_border, center_align, left_align)
        elif report_type == 'kpi':
            row = self._write_kpi_excel(ws, report_data, row, header_font, header_fill, cell_font, thin_border, center_align, left_align)

        # 调整列宽
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_type}_report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb.save(filepath)
        return filepath

    def _write_production_excel(self, ws, data, row, header_font, header_fill, cell_font, border, center_align, left_align):
        """写入生产报表内容"""
        summary = data.get('summary', {})

        # 写入汇总信息
        ws.cell(row=row, column=1, value="汇总信息").font = Font(bold=True, size=12)
        row += 1

        summary_items = [
            ('总计划数', summary.get('total_plans', 0)),
            ('已完成', summary.get('completed', 0)),
            ('进行中', summary.get('in_progress', 0)),
            ('未开始', summary.get('not_started', 0)),
            ('已延期', summary.get('delayed', 0)),
            ('完成率', f"{summary.get('completion_rate', 0):.1f}%"),
            ('延期率', f"{summary.get('delay_rate', 0):.1f}%"),
        ]

        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label).font = cell_font
            ws.cell(row=row, column=2, value=str(value)).font = cell_font
            row += 1

        row += 1

        # 写入明细表格
        details = data.get('details', [])
        if details:
            ws.cell(row=row, column=1, value="计划明细").font = Font(bold=True, size=12)
            row += 1

            headers = ['订单号', '客户名称', '产品名称', '开始日期', '结束日期', '状态', '进度']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_align
            row += 1

            for item in details:
                values = [
                    item.get('order_no', ''),
                    item.get('customer_name', ''),
                    item.get('product_name', ''),
                    item.get('start_date', ''),
                    item.get('end_date', ''),
                    item.get('status', ''),
                    f"{item.get('progress', 0)}%",
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = cell_font
                    cell.border = border
                    cell.alignment = center_align if col > 3 else left_align
                row += 1

        return row

    def _write_order_excel(self, ws, data, row, header_font, header_fill, cell_font, border, center_align, left_align):
        """写入订单报表内容"""
        summary = data.get('summary', {})

        # 汇总信息
        ws.cell(row=row, column=1, value="汇总信息").font = Font(bold=True, size=12)
        row += 1

        summary_items = [
            ('总订单数', summary.get('total_orders', 0)),
            ('客户数', summary.get('customer_count', 0)),
            ('已完成', summary.get('completed', 0)),
            ('进行中', summary.get('in_progress', 0)),
            ('准时交付率', f"{summary.get('on_time_rate', 0):.1f}%"),
        ]

        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label).font = cell_font
            ws.cell(row=row, column=2, value=str(value)).font = cell_font
            row += 1

        row += 1

        # 客户分布
        customer_stats = data.get('customer_stats', [])
        if customer_stats:
            ws.cell(row=row, column=1, value="客户订单分布").font = Font(bold=True, size=12)
            row += 1

            headers = ['客户名称', '订单数量', '占比']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            row += 1

            for item in customer_stats:
                values = [
                    item.get('customer_name', ''),
                    item.get('order_count', 0),
                    f"{item.get('percentage', 0):.1f}%",
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = cell_font
                    cell.border = border
                row += 1

        return row

    def _write_task_excel(self, ws, data, row, header_font, header_fill, cell_font, border, center_align, left_align):
        """写入任务报表内容"""
        summary = data.get('summary', {})

        # 汇总信息
        ws.cell(row=row, column=1, value="汇总信息").font = Font(bold=True, size=12)
        row += 1

        summary_items = [
            ('总任务数', summary.get('total_tasks', 0)),
            ('已完成', summary.get('completed', 0)),
            ('进行中', summary.get('in_progress', 0)),
            ('待处理', summary.get('pending', 0)),
            ('已逾期', summary.get('overdue', 0)),
            ('完成率', f"{summary.get('completion_rate', 0):.1f}%"),
            ('逾期率', f"{summary.get('overdue_rate', 0):.1f}%"),
        ]

        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label).font = cell_font
            ws.cell(row=row, column=2, value=str(value)).font = cell_font
            row += 1

        row += 1

        # 优先级分布
        priority_stats = data.get('priority_stats', [])
        if priority_stats:
            ws.cell(row=row, column=1, value="优先级分布").font = Font(bold=True, size=12)
            row += 1

            headers = ['优先级', '数量', '占比']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            row += 1

            for item in priority_stats:
                values = [
                    item.get('priority', ''),
                    item.get('count', 0),
                    f"{item.get('percentage', 0):.1f}%",
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = cell_font
                    cell.border = border
                row += 1

        return row

    def _write_kpi_excel(self, ws, data, row, header_font, header_fill, cell_font, border, center_align, left_align):
        """写入 KPI 报表内容"""
        summary = data.get('summary', {})

        # 综合评分
        ws.cell(row=row, column=1, value="综合评分").font = Font(bold=True, size=14)
        ws.cell(row=row, column=2, value=f"{summary.get('overall_score', 0):.1f}").font = Font(bold=True, size=14, color='2E7D32')
        row += 2

        # KPI 指标
        metrics = data.get('metrics', [])
        if metrics:
            ws.cell(row=row, column=1, value="KPI 指标详情").font = Font(bold=True, size=12)
            row += 1

            headers = ['指标名称', '当前值', '目标值', '得分', '权重', '加权得分']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_align
            row += 1

            for item in metrics:
                values = [
                    item.get('name', ''),
                    f"{item.get('value', 0):.1f}%",
                    f"{item.get('target', 0):.1f}%",
                    f"{item.get('score', 0):.1f}",
                    f"{item.get('weight', 0)*100:.0f}%",
                    f"{item.get('weighted_score', 0):.1f}",
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = cell_font
                    cell.border = border
                    cell.alignment = center_align
                row += 1

        return row

    def export_to_pdf(self, report_data, report_type, report_name=None):
        """
        导出为 PDF 文件

        Args:
            report_data: 报表数据字典
            report_type: 报表类型
            report_name: 报表名称（可选）

        Returns:
            str: 生成的文件路径
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab 未安装，无法生成 PDF 文件")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_type}_report_{timestamp}.pdf"
        filepath = os.path.join(self.output_dir, filename)

        # 创建文档
        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )

        # 准备内容
        elements = []
        styles = getSampleStyleSheet()

        # 尝试注册中文字体
        try:
            # Windows 系统字体路径
            font_paths = [
                'C:/Windows/Fonts/msyh.ttc',  # 微软雅黑
                'C:/Windows/Fonts/simsun.ttc',  # 宋体
            ]
            for font_path in font_paths:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('Chinese', font_path))
                    break
        except:
            pass

        # 标题
        title = report_name or self._get_report_title(report_type)
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,  # 居中
            spaceAfter=20
        )
        elements.append(Paragraph(title, title_style))

        # 日期范围
        if 'date_range' in report_data:
            dr = report_data['date_range']
            date_text = f"Date Range: {dr.get('start', '')} to {dr.get('end', '')}"
            elements.append(Paragraph(date_text, styles['Normal']))
            elements.append(Spacer(1, 10*mm))

        # 根据报表类型生成内容
        if report_type == 'production':
            elements.extend(self._build_production_pdf(report_data, styles))
        elif report_type == 'order':
            elements.extend(self._build_order_pdf(report_data, styles))
        elif report_type == 'task':
            elements.extend(self._build_task_pdf(report_data, styles))
        elif report_type == 'kpi':
            elements.extend(self._build_kpi_pdf(report_data, styles))

        # 生成 PDF
        doc.build(elements)
        return filepath

    def _build_production_pdf(self, data, styles):
        """构建生产报表 PDF 内容"""
        elements = []
        summary = data.get('summary', {})

        # 汇总表格
        elements.append(Paragraph("Summary", styles['Heading2']))

        summary_data = [
            ['Metric', 'Value'],
            ['Total Plans', str(summary.get('total_plans', 0))],
            ['Completed', str(summary.get('completed', 0))],
            ['In Progress', str(summary.get('in_progress', 0))],
            ['Delayed', str(summary.get('delayed', 0))],
            ['Completion Rate', f"{summary.get('completion_rate', 0):.1f}%"],
            ['Delay Rate', f"{summary.get('delay_rate', 0):.1f}%"],
        ]

        table = Table(summary_data, colWidths=[100, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10*mm))

        # 明细表格
        details = data.get('details', [])
        if details:
            elements.append(Paragraph("Plan Details", styles['Heading2']))

            detail_data = [['Order No', 'Customer', 'Product', 'Start', 'End', 'Status']]
            for item in details[:20]:  # 限制显示前20条
                detail_data.append([
                    item.get('order_no', ''),
                    item.get('customer_name', '')[:15],  # 截断
                    item.get('product_name', '')[:15],
                    item.get('start_date', ''),
                    item.get('end_date', ''),
                    item.get('status', ''),
                ])

            table = Table(detail_data, colWidths=[80, 80, 80, 60, 60, 60])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

        return elements

    def _build_order_pdf(self, data, styles):
        """构建订单报表 PDF 内容"""
        elements = []
        summary = data.get('summary', {})

        elements.append(Paragraph("Summary", styles['Heading2']))

        summary_data = [
            ['Metric', 'Value'],
            ['Total Orders', str(summary.get('total_orders', 0))],
            ['Customers', str(summary.get('customer_count', 0))],
            ['Completed', str(summary.get('completed', 0))],
            ['On-Time Rate', f"{summary.get('on_time_rate', 0):.1f}%"],
        ]

        table = Table(summary_data, colWidths=[100, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10*mm))

        # 客户分布
        customer_stats = data.get('customer_stats', [])
        if customer_stats:
            elements.append(Paragraph("Customer Distribution", styles['Heading2']))

            customer_data = [['Customer', 'Orders', 'Percentage']]
            for item in customer_stats[:10]:
                customer_data.append([
                    item.get('customer_name', '')[:20],
                    str(item.get('order_count', 0)),
                    f"{item.get('percentage', 0):.1f}%",
                ])

            table = Table(customer_data, colWidths=[150, 80, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

        return elements

    def _build_task_pdf(self, data, styles):
        """构建任务报表 PDF 内容"""
        elements = []
        summary = data.get('summary', {})

        elements.append(Paragraph("Summary", styles['Heading2']))

        summary_data = [
            ['Metric', 'Value'],
            ['Total Tasks', str(summary.get('total_tasks', 0))],
            ['Completed', str(summary.get('completed', 0))],
            ['In Progress', str(summary.get('in_progress', 0))],
            ['Overdue', str(summary.get('overdue', 0))],
            ['Completion Rate', f"{summary.get('completion_rate', 0):.1f}%"],
            ['Overdue Rate', f"{summary.get('overdue_rate', 0):.1f}%"],
        ]

        table = Table(summary_data, colWidths=[100, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)

        return elements

    def _build_kpi_pdf(self, data, styles):
        """构建 KPI 报表 PDF 内容"""
        elements = []
        summary = data.get('summary', {})

        # 综合评分
        score_style = ParagraphStyle(
            'Score',
            parent=styles['Heading1'],
            fontSize=24,
            alignment=1,
            textColor=colors.HexColor('#2E7D32')
        )
        elements.append(Paragraph(f"Overall Score: {summary.get('overall_score', 0):.1f}", score_style))
        elements.append(Spacer(1, 10*mm))

        # KPI 指标表格
        metrics = data.get('metrics', [])
        if metrics:
            elements.append(Paragraph("KPI Metrics", styles['Heading2']))

            kpi_data = [['Metric', 'Value', 'Target', 'Score', 'Weight', 'Weighted']]
            for item in metrics:
                kpi_data.append([
                    item.get('name', ''),
                    f"{item.get('value', 0):.1f}%",
                    f"{item.get('target', 0):.1f}%",
                    f"{item.get('score', 0):.1f}",
                    f"{item.get('weight', 0)*100:.0f}%",
                    f"{item.get('weighted_score', 0):.1f}",
                ])

            table = Table(kpi_data, colWidths=[100, 60, 60, 60, 60, 60])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

        return elements

    def export_to_csv(self, report_data, report_type, report_name=None):
        """
        导出为 CSV 文件

        Args:
            report_data: 报表数据字典
            report_type: 报表类型
            report_name: 报表名称（可选）

        Returns:
            str: 生成的文件路径
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_type}_report_{timestamp}.csv"
        filepath = os.path.join(self.output_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # 根据报表类型写入不同内容
            if report_type == 'production':
                self._write_production_csv(writer, report_data)
            elif report_type == 'order':
                self._write_order_csv(writer, report_data)
            elif report_type == 'task':
                self._write_task_csv(writer, report_data)
            elif report_type == 'kpi':
                self._write_kpi_csv(writer, report_data)

        return filepath

    def _write_production_csv(self, writer, data):
        """写入生产报表 CSV"""
        summary = data.get('summary', {})

        # 汇总
        writer.writerow(['Summary'])
        writer.writerow(['Total Plans', summary.get('total_plans', 0)])
        writer.writerow(['Completed', summary.get('completed', 0)])
        writer.writerow(['In Progress', summary.get('in_progress', 0)])
        writer.writerow(['Delayed', summary.get('delayed', 0)])
        writer.writerow(['Completion Rate', f"{summary.get('completion_rate', 0):.1f}%"])
        writer.writerow([])

        # 明细
        details = data.get('details', [])
        if details:
            writer.writerow(['Details'])
            writer.writerow(['Order No', 'Customer', 'Product', 'Start Date', 'End Date', 'Status', 'Progress'])
            for item in details:
                writer.writerow([
                    item.get('order_no', ''),
                    item.get('customer_name', ''),
                    item.get('product_name', ''),
                    item.get('start_date', ''),
                    item.get('end_date', ''),
                    item.get('status', ''),
                    f"{item.get('progress', 0)}%",
                ])

    def _write_order_csv(self, writer, data):
        """写入订单报表 CSV"""
        summary = data.get('summary', {})

        writer.writerow(['Summary'])
        writer.writerow(['Total Orders', summary.get('total_orders', 0)])
        writer.writerow(['Customers', summary.get('customer_count', 0)])
        writer.writerow(['Completed', summary.get('completed', 0)])
        writer.writerow(['On-Time Rate', f"{summary.get('on_time_rate', 0):.1f}%"])
        writer.writerow([])

        customer_stats = data.get('customer_stats', [])
        if customer_stats:
            writer.writerow(['Customer Distribution'])
            writer.writerow(['Customer', 'Orders', 'Percentage'])
            for item in customer_stats:
                writer.writerow([
                    item.get('customer_name', ''),
                    item.get('order_count', 0),
                    f"{item.get('percentage', 0):.1f}%",
                ])

    def _write_task_csv(self, writer, data):
        """写入任务报表 CSV"""
        summary = data.get('summary', {})

        writer.writerow(['Summary'])
        writer.writerow(['Total Tasks', summary.get('total_tasks', 0)])
        writer.writerow(['Completed', summary.get('completed', 0)])
        writer.writerow(['In Progress', summary.get('in_progress', 0)])
        writer.writerow(['Overdue', summary.get('overdue', 0)])
        writer.writerow(['Completion Rate', f"{summary.get('completion_rate', 0):.1f}%"])
        writer.writerow(['Overdue Rate', f"{summary.get('overdue_rate', 0):.1f}%"])
        writer.writerow([])

        details = data.get('details', [])
        if details:
            writer.writerow(['Details'])
            writer.writerow(['Title', 'Priority', 'Status', 'Due Date', 'Assignee'])
            for item in details:
                writer.writerow([
                    item.get('title', ''),
                    item.get('priority', ''),
                    item.get('status', ''),
                    item.get('due_date', ''),
                    item.get('assignee_name', ''),
                ])

    def _write_kpi_csv(self, writer, data):
        """写入 KPI 报表 CSV"""
        summary = data.get('summary', {})

        writer.writerow(['Overall Score', f"{summary.get('overall_score', 0):.1f}"])
        writer.writerow([])

        metrics = data.get('metrics', [])
        if metrics:
            writer.writerow(['KPI Metrics'])
            writer.writerow(['Metric', 'Value', 'Target', 'Score', 'Weight', 'Weighted Score'])
            for item in metrics:
                writer.writerow([
                    item.get('name', ''),
                    f"{item.get('value', 0):.1f}%",
                    f"{item.get('target', 0):.1f}%",
                    f"{item.get('score', 0):.1f}",
                    f"{item.get('weight', 0)*100:.0f}%",
                    f"{item.get('weighted_score', 0):.1f}",
                ])

    def _get_report_title(self, report_type):
        """获取报表标题"""
        titles = {
            'production': 'Production Report',
            'order': 'Order Report',
            'task': 'Task Report',
            'kpi': 'KPI Report',
            'custom': 'Custom Report',
        }
        return titles.get(report_type, 'Report')

    def get_file_size(self, filepath):
        """获取文件大小（字节）"""
        if os.path.exists(filepath):
            return os.path.getsize(filepath)
        return 0
