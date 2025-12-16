# -*- coding: utf-8 -*-
"""
订单导入导出 API
提供 Excel 批量导入、导出功能
"""
from __future__ import annotations

import io
import os
from datetime import datetime, date
from decimal import Decimal
from flask import jsonify, request, send_file, current_app
from werkzeug.utils import secure_filename

from . import orders  # 复用 orders.py 中的 bp 和工具函数
from ..models.core import Order, OrderLine, OrderStatus
from ..models.customer import Customer
from .. import db

bp = orders.bp  # 使用同一个 Blueprint


# === 导入模板字段定义 ===
IMPORT_TEMPLATE_COLUMNS = [
    ("order_no", "订单号", True),
    ("customer_code", "客户编码", True),
    ("product", "内部图号", True),
    ("cn_name", "中文名称", False),
    ("order_qty", "订单数量", True),
    ("unit_price", "单价", False),
    ("currency", "币种", False),
    ("due_date", "交货日期", False),
    ("process_content", "加工内容", False),
    ("packing_req", "包装要求", False),
    ("remark", "备注", False),
]


def _parse_date(value):
    """解析日期值（支持多种格式）"""
    if not value:
        return None
    if isinstance(value, (date, datetime)):
        return value if isinstance(value, date) else value.date()
    try:
        # 尝试常见格式
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%d/%m/%Y', '%d-%m-%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except:
                continue
        # Excel 日期数值
        if isinstance(value, (int, float)):
            from datetime import timedelta
            base_date = datetime(1899, 12, 30)
            return (base_date + timedelta(days=int(value))).date()
    except:
        pass
    return None


def _safe_str(value):
    """安全转换为字符串"""
    if value is None:
        return None
    return str(value).strip()


def _safe_int(value):
    """安全转换为整数"""
    if value is None or value == '':
        return 0
    try:
        return int(float(str(value).replace(',', '')))
    except:
        return 0


def _safe_decimal(value):
    """安全转换为 Decimal"""
    if value is None or value == '':
        return Decimal('0')
    try:
        return Decimal(str(value).replace(',', ''))
    except:
        return Decimal('0')


# === 导出订单 API ===

@bp.route('/export', methods=['GET'])
def export_orders():
    """
    导出订单到 Excel
    参数:
      - status: 状态筛选
      - date_from: 开始日期
      - date_to: 结束日期
      - customer_id: 客户ID
      - keyword: 关键词
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # 构建查询
        query = Order.query

        # 状态筛选
        status = request.args.get('status')
        if status:
            query = query.filter(Order.status == status)

        # 日期筛选
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        if date_from:
            query = query.filter(Order.order_date >= date_from)
        if date_to:
            query = query.filter(Order.order_date <= date_to)

        # 客户筛选
        customer_id = request.args.get('customer_id')
        if customer_id:
            query = query.filter(Order.customer_id == int(customer_id))

        # 关键词搜索
        keyword = request.args.get('keyword', '').strip()
        if keyword:
            from sqlalchemy import or_
            query = query.filter(or_(
                Order.order_no.like(f'%{keyword}%'),
                Order.product.like(f'%{keyword}%'),
                Order.customer_code.like(f'%{keyword}%')
            ))

        orders_list = query.order_by(Order.id.desc()).all()

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "订单列表"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 定义导出列
        headers = [
            "订单号", "客户编码", "内部图号", "中文名称", "订单数量",
            "已交数量", "欠交数量", "单价", "币种", "交货日期",
            "状态", "加工内容", "包装要求", "备注", "下单日期"
        ]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 状态映射
        status_map = {
            OrderStatus.DRAFT.value: "草稿",
            OrderStatus.PENDING.value: "待审批",
            OrderStatus.CONFIRMED.value: "已确认",
            OrderStatus.IN_PRODUCTION.value: "生产中",
            OrderStatus.IN_DELIVERY.value: "交货中",
            OrderStatus.COMPLETED.value: "已完成",
            OrderStatus.CANCELLED.value: "已取消",
        }

        # 写入数据
        for row, order in enumerate(orders_list, 2):
            ws.cell(row=row, column=1, value=order.order_no)
            ws.cell(row=row, column=2, value=order.customer_code)
            ws.cell(row=row, column=3, value=order.product)
            ws.cell(row=row, column=4, value=order.cn_name)
            ws.cell(row=row, column=5, value=order.order_qty or order.qty_total or 0)
            ws.cell(row=row, column=6, value=order.delivered_qty or 0)
            ws.cell(row=row, column=7, value=order.deficit_qty or 0)
            ws.cell(row=row, column=8, value=float(order.unit_price) if order.unit_price else 0)
            ws.cell(row=row, column=9, value=order.currency or 'CNY')
            ws.cell(row=row, column=10, value=order.due_date.strftime('%Y-%m-%d') if order.due_date else '')
            ws.cell(row=row, column=11, value=status_map.get(order.status, order.status))
            ws.cell(row=row, column=12, value=order.process_content)
            ws.cell(row=row, column=13, value=order.packing_req)
            ws.cell(row=row, column=14, value=order.remark)
            ws.cell(row=row, column=15, value=order.order_date.strftime('%Y-%m-%d') if order.order_date else '')

            # 添加边框
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border

        # 调整列宽
        column_widths = [15, 15, 20, 25, 12, 12, 12, 12, 8, 12, 10, 20, 20, 30, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        filename = f"订单导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# === 下载导入模板 API ===

@bp.route('/export/template', methods=['GET'])
def download_import_template():
    """下载订单导入模板"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "订单导入模板"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        required_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col, (field, label, required) in enumerate(IMPORT_TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=f"{label}{'*' if required else ''}")
            cell.font = header_font
            cell.fill = required_fill if required else header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入示例数据
        example_data = [
            ("ORD-20250101-001", "CUST001", "PART-001", "产品A", 100, 10.5, "CNY", "2025-02-01", "精加工", "纸箱包装", "测试订单"),
            ("ORD-20250101-002", "CUST002", "PART-002", "产品B", 200, 20.0, "USD", "2025-02-15", "组装", "木箱包装", ""),
        ]

        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border

        # 调整列宽
        column_widths = [20, 15, 20, 20, 12, 12, 8, 12, 15, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 添加说明 sheet
        ws_help = wb.create_sheet(title="填写说明")
        instructions = [
            ["字段说明"],
            [""],
            ["字段名", "是否必填", "说明"],
            ["订单号", "是", "订单的唯一标识，不能重复"],
            ["客户编码", "是", "客户的编码，需要在系统中已存在"],
            ["内部图号", "是", "产品的内部图号"],
            ["中文名称", "否", "产品的中文名称"],
            ["订单数量", "是", "订单数量，必须为正整数"],
            ["单价", "否", "产品单价，默认为0"],
            ["币种", "否", "货币类型，默认为CNY。支持: CNY, USD, EUR, JPY"],
            ["交货日期", "否", "要求交货日期，格式: YYYY-MM-DD"],
            ["加工内容", "否", "加工内容描述"],
            ["包装要求", "否", "包装要求描述"],
            ["备注", "否", "其他备注信息"],
        ]

        for row, data in enumerate(instructions, 1):
            for col, value in enumerate(data, 1):
                ws_help.cell(row=row, column=col, value=value)

        # 调整说明 sheet 列宽
        ws_help.column_dimensions['A'].width = 15
        ws_help.column_dimensions['B'].width = 10
        ws_help.column_dimensions['C'].width = 50

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='订单导入模板.xlsx'
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# === 预览导入数据 API ===

@bp.route('/import/preview', methods=['POST'])
def preview_import_orders():
    """
    预览导入数据（不实际导入）
    """
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "请上传文件"}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({"success": False, "error": "文件名为空"}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "error": "仅支持 Excel 文件 (.xlsx, .xls)"}), 400

    try:
        from openpyxl import load_workbook

        # 读取 Excel
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # 解析表头
        headers = [cell.value for cell in ws[1]]

        # 字段映射
        field_map = {label: field for field, label, _ in IMPORT_TEMPLATE_COLUMNS}
        field_map.update({f"{label}*": field for field, label, required in IMPORT_TEMPLATE_COLUMNS if required})

        column_mapping = {}
        for col, header in enumerate(headers):
            if header:
                field = field_map.get(str(header).strip())
                if field:
                    column_mapping[col] = field

        # 解析数据
        preview_data = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # 跳过空行
                continue

            row_data = {}
            for col, field in column_mapping.items():
                if col < len(row):
                    row_data[field] = row[col]

            # 验证必填字段
            row_errors = []
            for field, label, required in IMPORT_TEMPLATE_COLUMNS:
                if required and not row_data.get(field):
                    row_errors.append(f"缺少必填字段: {label}")

            # 验证订单号唯一性
            order_no = _safe_str(row_data.get('order_no'))
            if order_no:
                existing = Order.query.filter_by(order_no=order_no).first()
                if existing:
                    row_errors.append(f"订单号已存在: {order_no}")

            # 验证客户编码
            customer_code = _safe_str(row_data.get('customer_code'))
            if customer_code:
                customer = Customer.query.filter_by(code=customer_code).first()
                if not customer:
                    row_errors.append(f"客户编码不存在: {customer_code}")
                else:
                    row_data['customer_id'] = customer.id
                    row_data['customer_name'] = customer.short_name or customer.name

            # 验证数量
            order_qty = _safe_int(row_data.get('order_qty'))
            if order_qty <= 0:
                row_errors.append("订单数量必须大于0")

            preview_item = {
                "row": row_idx,
                "data": {
                    "order_no": order_no,
                    "customer_code": customer_code,
                    "customer_name": row_data.get('customer_name'),
                    "product": _safe_str(row_data.get('product')),
                    "cn_name": _safe_str(row_data.get('cn_name')),
                    "order_qty": order_qty,
                    "unit_price": float(_safe_decimal(row_data.get('unit_price'))),
                    "currency": _safe_str(row_data.get('currency')) or 'CNY',
                    "due_date": _parse_date(row_data.get('due_date')).isoformat() if _parse_date(row_data.get('due_date')) else None,
                    "process_content": _safe_str(row_data.get('process_content')),
                    "packing_req": _safe_str(row_data.get('packing_req')),
                    "remark": _safe_str(row_data.get('remark')),
                },
                "valid": len(row_errors) == 0,
                "errors": row_errors
            }

            preview_data.append(preview_item)
            if row_errors:
                errors.append({"row": row_idx, "errors": row_errors})

        valid_count = sum(1 for item in preview_data if item['valid'])
        invalid_count = len(preview_data) - valid_count

        return jsonify({
            "success": True,
            "total": len(preview_data),
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "preview": preview_data,
            "errors": errors
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# === 批量导入订单 API ===

@bp.route('/import', methods=['POST'])
def import_orders():
    """
    批量导入订单
    """
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "请上传文件"}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({"success": False, "error": "文件名为空"}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "error": "仅支持 Excel 文件 (.xlsx, .xls)"}), 400

    skip_errors = request.form.get('skip_errors', 'false').lower() == 'true'

    try:
        from openpyxl import load_workbook

        # 读取 Excel
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # 解析表头
        headers = [cell.value for cell in ws[1]]

        # 字段映射
        field_map = {label: field for field, label, _ in IMPORT_TEMPLATE_COLUMNS}
        field_map.update({f"{label}*": field for field, label, required in IMPORT_TEMPLATE_COLUMNS if required})

        column_mapping = {}
        for col, header in enumerate(headers):
            if header:
                field = field_map.get(str(header).strip())
                if field:
                    column_mapping[col] = field

        # 导入结果统计
        success_count = 0
        error_count = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # 跳过空行
                continue

            row_data = {}
            for col, field in column_mapping.items():
                if col < len(row):
                    row_data[field] = row[col]

            # 验证必填字段
            row_errors = []
            for field, label, required in IMPORT_TEMPLATE_COLUMNS:
                if required and not row_data.get(field):
                    row_errors.append(f"缺少必填字段: {label}")

            if row_errors:
                error_count += 1
                errors.append({"row": row_idx, "errors": row_errors})
                if not skip_errors:
                    continue

            # 验证订单号唯一性
            order_no = _safe_str(row_data.get('order_no'))
            if order_no:
                existing = Order.query.filter_by(order_no=order_no).first()
                if existing:
                    error_count += 1
                    errors.append({"row": row_idx, "errors": [f"订单号已存在: {order_no}"]})
                    continue

            # 验证客户编码
            customer_code = _safe_str(row_data.get('customer_code'))
            customer = None
            if customer_code:
                customer = Customer.query.filter_by(code=customer_code).first()
                if not customer:
                    error_count += 1
                    errors.append({"row": row_idx, "errors": [f"客户编码不存在: {customer_code}"]})
                    continue

            # 验证数量
            order_qty = _safe_int(row_data.get('order_qty'))
            if order_qty <= 0:
                error_count += 1
                errors.append({"row": row_idx, "errors": ["订单数量必须大于0"]})
                continue

            try:
                # 创建订单
                order = Order(
                    order_no=order_no,
                    customer_id=customer.id if customer else None,
                    customer_code=customer_code,
                    product=_safe_str(row_data.get('product')),
                    cn_name=_safe_str(row_data.get('cn_name')),
                    order_qty=order_qty,
                    qty_total=order_qty,
                    unit_price=_safe_decimal(row_data.get('unit_price')),
                    currency=_safe_str(row_data.get('currency')) or 'CNY',
                    due_date=_parse_date(row_data.get('due_date')),
                    process_content=_safe_str(row_data.get('process_content')),
                    packing_req=_safe_str(row_data.get('packing_req')),
                    remark=_safe_str(row_data.get('remark')),
                    status=OrderStatus.DRAFT.value,
                    order_date=date.today(),
                    delivered_qty=0,
                    deficit_qty=order_qty
                )

                db.session.add(order)
                db.session.flush()  # 获取 order.id

                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append({"row": row_idx, "errors": [str(e)]})
                continue

        # 提交事务
        if success_count > 0:
            db.session.commit()

        return jsonify({
            "success": True,
            "message": f"导入完成：成功 {success_count} 条，失败 {error_count} 条",
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors[:50]  # 最多返回前50条错误
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
