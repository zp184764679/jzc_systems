"""
导出服务 - 支持 PDF 和 Excel 导出
"""
import io
from datetime import datetime
from typing import List, Dict, Any, Optional

# PDF 导出
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Excel 导出
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 注册中文字体 (使用系统字体)
import os
FONT_PATHS = [
    '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',
    'C:/Windows/Fonts/msyh.ttc',  # Windows 微软雅黑
    'C:/Windows/Fonts/simhei.ttf',  # Windows 黑体
]

CHINESE_FONT = None
for font_path in FONT_PATHS:
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
            CHINESE_FONT = 'ChineseFont'
            break
        except:
            continue

if not CHINESE_FONT:
    CHINESE_FONT = 'Helvetica'  # 回退到默认字体


class ExportService:
    """导出服务"""

    # 状态映射
    STATUS_MAP = {
        'planning': '规划中',
        'in_progress': '进行中',
        'on_hold': '暂停',
        'completed': '已完成',
        'cancelled': '已取消',
        'pending': '待开始',
        'blocked': '受阻',
    }

    # 优先级映射
    PRIORITY_MAP = {
        'urgent': '紧急',
        'high': '高',
        'normal': '普通',
        'low': '低',
    }

    @staticmethod
    def get_status_text(status: str) -> str:
        return ExportService.STATUS_MAP.get(status, status or '未知')

    @staticmethod
    def get_priority_text(priority: str) -> str:
        return ExportService.PRIORITY_MAP.get(priority, priority or '普通')

    @staticmethod
    def format_date(date_value) -> str:
        if not date_value:
            return '-'
        if isinstance(date_value, str):
            return date_value[:10] if len(date_value) >= 10 else date_value
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')
        return str(date_value)

    # ============ Excel 导出 ============

    @staticmethod
    def export_projects_to_excel(projects: List[Dict]) -> bytes:
        """导出项目列表到 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "项目列表"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = ['项目编号', '项目名称', '客户', '部件番号', '订单号', '状态', '优先级', '进度', '计划开始', '计划结束', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, project in enumerate(projects, 2):
            data = [
                project.get('project_no', ''),
                project.get('name', ''),
                project.get('customer', ''),
                project.get('part_number', ''),
                project.get('order_no', ''),
                ExportService.get_status_text(project.get('status')),
                ExportService.get_priority_text(project.get('priority')),
                f"{project.get('progress_percentage', 0)}%",
                ExportService.format_date(project.get('planned_start_date')),
                ExportService.format_date(project.get('planned_end_date')),
                ExportService.format_date(project.get('created_at')),
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 调整列宽
        column_widths = [15, 30, 20, 15, 15, 10, 10, 10, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_tasks_to_excel(tasks: List[Dict], project_name: str = None) -> bytes:
        """导出任务列表到 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "任务列表"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="34C759", end_color="34C759", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 如果有项目名称，添加标题行
        start_row = 1
        if project_name:
            ws.cell(row=1, column=1, value=f"项目: {project_name}")
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.merge_cells('A1:H1')
            start_row = 3

        # 表头
        headers = ['任务编号', '任务标题', '状态', '优先级', '完成度', '负责人', '截止日期', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, task in enumerate(tasks, start_row + 1):
            data = [
                task.get('task_no', ''),
                task.get('title', ''),
                ExportService.get_status_text(task.get('status')),
                ExportService.get_priority_text(task.get('priority')),
                f"{task.get('completion_percentage', 0)}%",
                task.get('assigned_to_name', '') or '-',
                ExportService.format_date(task.get('due_date')),
                ExportService.format_date(task.get('created_at')),
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 调整列宽
        column_widths = [15, 40, 10, 10, 10, 15, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 冻结首行
        ws.freeze_panes = f'A{start_row + 1}'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ============ PDF 导出 ============

    @staticmethod
    def export_project_report_to_pdf(project: Dict, tasks: List[Dict], phases: List[Dict] = None) -> bytes:
        """导出单个项目报告到 PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )

        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ChineseTitle',
            parent=styles['Title'],
            fontName=CHINESE_FONT,
            fontSize=18,
            spaceAfter=20
        )
        heading_style = ParagraphStyle(
            'ChineseHeading',
            parent=styles['Heading2'],
            fontName=CHINESE_FONT,
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10
        )
        normal_style = ParagraphStyle(
            'ChineseNormal',
            parent=styles['Normal'],
            fontName=CHINESE_FONT,
            fontSize=10,
            leading=14
        )

        elements = []

        # 标题
        elements.append(Paragraph(f"项目报告: {project.get('name', '')}", title_style))
        elements.append(Paragraph(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}", normal_style))
        elements.append(Spacer(1, 10*mm))

        # 项目基本信息
        elements.append(Paragraph("项目概况", heading_style))
        info_data = [
            ['项目编号', project.get('project_no', '-'), '状态', ExportService.get_status_text(project.get('status'))],
            ['客户', project.get('customer', '-'), '优先级', ExportService.get_priority_text(project.get('priority'))],
            ['部件番号', project.get('part_number', '-'), '进度', f"{project.get('progress_percentage', 0)}%"],
            ['订单号', project.get('order_no', '-'), '创建时间', ExportService.format_date(project.get('created_at'))],
            ['计划开始', ExportService.format_date(project.get('planned_start_date')), '计划结束', ExportService.format_date(project.get('planned_end_date'))],
        ]

        info_table = Table(info_data, colWidths=[60*mm, 60*mm, 40*mm, 40*mm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.95, 0.95, 0.97)),
            ('BACKGROUND', (2, 0), (2, -1), colors.Color(0.95, 0.95, 0.97)),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10*mm))

        # 项目描述
        if project.get('description'):
            elements.append(Paragraph("项目描述", heading_style))
            elements.append(Paragraph(project.get('description', ''), normal_style))
            elements.append(Spacer(1, 5*mm))

        # 任务列表
        if tasks:
            elements.append(Paragraph(f"任务列表 ({len(tasks)} 个任务)", heading_style))

            task_headers = ['任务编号', '任务标题', '状态', '完成度', '负责人', '截止日期']
            task_data = [task_headers]

            for task in tasks:
                task_data.append([
                    task.get('task_no', '-'),
                    task.get('title', '-')[:20] + ('...' if len(task.get('title', '')) > 20 else ''),
                    ExportService.get_status_text(task.get('status')),
                    f"{task.get('completion_percentage', 0)}%",
                    task.get('assigned_to_name', '-') or '-',
                    ExportService.format_date(task.get('due_date')),
                ])

            task_table = Table(task_data, colWidths=[30*mm, 50*mm, 25*mm, 20*mm, 30*mm, 25*mm])
            task_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.48, 1.0)),  # 蓝色表头
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (2, 0), (3, -1), 'CENTER'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                # 交替行颜色
                *[('BACKGROUND', (0, i), (-1, i), colors.Color(0.97, 0.97, 0.99)) for i in range(2, len(task_data), 2)],
            ]))
            elements.append(task_table)

        # 构建 PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_projects_summary_to_pdf(projects: List[Dict]) -> bytes:
        """导出项目汇总报告到 PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),  # 横向
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=15*mm,
            bottomMargin=15*mm
        )

        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ChineseTitle',
            parent=styles['Title'],
            fontName=CHINESE_FONT,
            fontSize=16,
            spaceAfter=15
        )
        normal_style = ParagraphStyle(
            'ChineseNormal',
            parent=styles['Normal'],
            fontName=CHINESE_FONT,
            fontSize=9
        )

        elements = []

        # 标题
        elements.append(Paragraph("项目汇总报告", title_style))
        elements.append(Paragraph(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  共 {len(projects)} 个项目", normal_style))
        elements.append(Spacer(1, 8*mm))

        # 统计摘要
        status_counts = {}
        for p in projects:
            status = p.get('status', 'unknown')
            status_counts[status] = status_counts.get(status, 0) + 1

        summary_text = "  |  ".join([
            f"{ExportService.get_status_text(k)}: {v}" for k, v in status_counts.items()
        ])
        elements.append(Paragraph(f"状态统计: {summary_text}", normal_style))
        elements.append(Spacer(1, 5*mm))

        # 项目表格
        headers = ['项目编号', '项目名称', '客户', '部件番号', '状态', '优先级', '进度', '计划开始', '计划结束']
        table_data = [headers]

        for project in projects:
            table_data.append([
                project.get('project_no', '-'),
                project.get('name', '-')[:25] + ('...' if len(project.get('name', '')) > 25 else ''),
                project.get('customer', '-')[:15] + ('...' if len(project.get('customer', '')) > 15 else ''),
                project.get('part_number', '-'),
                ExportService.get_status_text(project.get('status')),
                ExportService.get_priority_text(project.get('priority')),
                f"{project.get('progress_percentage', 0)}%",
                ExportService.format_date(project.get('planned_start_date')),
                ExportService.format_date(project.get('planned_end_date')),
            ])

        col_widths = [28*mm, 55*mm, 40*mm, 28*mm, 22*mm, 18*mm, 18*mm, 25*mm, 25*mm]
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.48, 1.0)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (4, 0), (6, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            # 交替行颜色
            *[('BACKGROUND', (0, i), (-1, i), colors.Color(0.97, 0.97, 0.99)) for i in range(2, len(table_data), 2)],
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
