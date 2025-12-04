# services/quote_document_generator.py
"""
报价单文档生成服务
支持PDF和Excel格式导出
"""
import os
import logging
from typing import Dict, Optional, List
from datetime import datetime
from io import BytesIO
import xlrd
from xlwt import Workbook, XFStyle, Font, Alignment, Borders, Pattern, easyxf
from xlutils.copy import copy as xl_copy

logger = logging.getLogger(__name__)


class QuoteDocumentGenerator:
    """报价单文档生成器"""

    def __init__(self):
        self.company_name = "永州雅力德精密零件加工厂"
        self.company_address = "永州市雅力德工业园"
        self.company_contact = "联系电话: 0746-1234567"

        # 晨龙精密报价单模板路径
        self.template_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'templates',
            'quote_template.xls'
        )

        # 初始化样式
        self._init_styles()

    def _init_styles(self):
        """初始化单元格样式 - 使用easyxf确保样式一致性"""
        # 基础边框定义
        border_str = 'border: left thin, right thin, top thin, bottom thin;'
        font_str = 'font: name 宋体, height 200;'  # 10pt

        # 表头样式（TO:, ATTN:, 报价单号等）- 无边框
        self.header_style = easyxf('font: name 宋体, height 220; align: vert center;')

        # 日期/电话/传真样式（右侧）- 无边框
        self.header_right_style = easyxf('font: name 宋体, height 220; align: vert center;')

        # 序号样式（居中，有边框）
        self.index_style = easyxf(
            f'{font_str} align: horiz center, vert center; {border_str}'
        )

        # 文本样式（左对齐，有边框）
        self.text_style = easyxf(
            f'{font_str} align: horiz left, vert center; {border_str}'
        )

        # 文本样式-居中（有边框）
        self.text_center_style = easyxf(
            f'{font_str} align: horiz center, vert center; {border_str}'
        )

        # 数字样式（居中，有边框）
        self.number_style = easyxf(
            f'{font_str} align: horiz center, vert center; {border_str}',
            num_format_str='0.0'
        )

        # 价格样式（居中，保留4位小数，有边框）
        self.price_style = easyxf(
            f'{font_str} align: horiz center, vert center; {border_str}',
            num_format_str='0.0000'
        )

        # 数量样式（居中，整数带千分位，有边框）
        self.qty_style = easyxf(
            f'{font_str} align: horiz center, vert center; {border_str}',
            num_format_str='#,##0'
        )

    def generate_excel(self, quote_data: Dict) -> BytesIO:
        """
        生成Excel报价单

        Args:
            quote_data: 报价数据

        Returns:
            Excel文件的BytesIO对象
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "报价单"

            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30

            # 标题
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = "机加工零件报价单"
            title_cell.font = Font(size=18, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 公司信息
            row = 3
            ws[f'A{row}'] = self.company_name
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = self.company_address
            row += 1
            ws[f'A{row}'] = self.company_contact
            row += 2

            # 基本信息
            quote = quote_data.get('quote', {})
            drawing_info = quote_data.get('drawing_info', {})
            material_info = quote_data.get('material', {})

            ws[f'A{row}'] = "报价单号:"
            ws[f'B{row}'] = quote_data.get('quote_number', 'N/A')
            ws[f'C{row}'] = "日期:"
            ws[f'D{row}'] = datetime.now().strftime('%Y-%m-%d')
            row += 1

            ws[f'A{row}'] = "客户名称:"
            ws[f'B{row}'] = quote_data.get('customer_name', 'N/A')
            row += 1

            ws[f'A{row}'] = "图号:"
            ws[f'B{row}'] = drawing_info.get('drawing_number', 'N/A')
            row += 1

            ws[f'A{row}'] = "产品名称:"
            ws[f'B{row}'] = drawing_info.get('product_name', 'N/A')
            row += 1

            ws[f'A{row}'] = "材质:"
            ws[f'B{row}'] = drawing_info.get('material', 'N/A')
            row += 2

            # 报价明细表头
            headers = ['项目', '规格', '单价(元)', '备注']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            row += 1

            # 明细行
            details = [
                ("材料费", f"密度: {material_info.get('details', {}).get('weight_kg', 0):.4f}kg",
                 quote.get('material_cost', 0), "含不良率和材管率"),
                ("加工费", f"工序数: {len(quote_data.get('process', {}).get('process_details', []))}",
                 quote.get('process_cost', 0), "含各工序成本"),
                ("管理费", f"{quote.get('rates', {}).get('management_rate', 0.045)*100:.1f}%",
                 quote.get('management_cost', 0), ""),
                ("其他费用", "包装+消耗品", quote.get('other_cost', 0), ""),
                ("利润", f"{quote.get('rates', {}).get('profit_rate', 0.15)*100:.1f}%",
                 quote.get('profit', 0), ""),
            ]

            for item, spec, price, remark in details:
                ws[f'A{row}'] = item
                ws[f'B{row}'] = spec
                ws[f'C{row}'] = f"{price:.4f}"
                ws[f'D{row}'] = remark

                # 添加边框
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                row += 1

            # 总计
            ws[f'A{row}'] = "单价总计"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'C{row}'] = f"{quote.get('total_price', 0):.4f}"
            ws[f'C{row}'].font = Font(bold=True, size=12, color="FF0000")
            row += 2

            # 批量信息
            ws[f'A{row}'] = f"批量: {quote_data.get('lot_size', 2000)} 件"
            row += 1
            ws[f'A{row}'] = f"总金额: ¥{quote.get('total_price', 0) * quote_data.get('lot_size', 2000):,.2f}"
            ws[f'A{row}'].font = Font(bold=True, size=14, color="FF0000")
            row += 3

            # 备注
            ws[f'A{row}'] = "备注:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = "1. 报价有效期30天"
            row += 1
            ws[f'A{row}'] = "2. 价格含税，不含运费"
            row += 1
            ws[f'A{row}'] = "3. 交货周期另议"

            # 保存到BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info("Excel报价单生成成功")
            return output

        except Exception as e:
            logger.error(f"生成Excel报价单失败: {e}")
            raise ValueError(f"生成Excel失败: {e}")

    def generate_pdf(self, quote_data: Dict) -> BytesIO:
        """
        生成PDF报价单

        Args:
            quote_data: 报价数据

        Returns:
            PDF文件的BytesIO对象
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import Table, TableStyle

            # 注册中文字体（使用系统字体）
            try:
                font_path = r"C:\Windows\Fonts\msyh.ttc"
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                    font_name = 'CustomFont'
                else:
                    font_name = 'Helvetica'
            except:
                font_name = 'Helvetica'

            # 创建PDF
            output = BytesIO()
            c = canvas.Canvas(output, pagesize=A4)
            width, height = A4

            # 提取数据
            quote = quote_data.get('quote', {})
            drawing_info = quote_data.get('drawing_info', {})
            material_info = quote_data.get('material', {})

            # 标题
            c.setFont(font_name if font_name == 'Helvetica' else font_name, 20)
            c.drawCentredString(width/2, height - 50*mm, "机加工零件报价单")

            # 公司信息
            y = height - 60*mm
            c.setFont(font_name if font_name == 'Helvetica' else font_name, 10)
            c.drawString(50*mm, y, self.company_name)
            y -= 15
            c.drawString(50*mm, y, self.company_address)
            y -= 15
            c.drawString(50*mm, y, self.company_contact)
            y -= 30

            # 基本信息
            c.drawString(30*mm, y, f"报价单号: {quote_data.get('quote_number', 'N/A')}")
            c.drawString(120*mm, y, f"日期: {datetime.now().strftime('%Y-%m-%d')}")
            y -= 15

            c.drawString(30*mm, y, f"客户名称: {quote_data.get('customer_name', 'N/A')}")
            y -= 15

            c.drawString(30*mm, y, f"图号: {drawing_info.get('drawing_number', 'N/A')}")
            y -= 15

            c.drawString(30*mm, y, f"材质: {drawing_info.get('material', 'N/A')}")
            y -= 30

            # 表格数据
            table_data = [
                ['项目', '规格', '单价(元)', '备注'],
                ['材料费', f"{material_info.get('details', {}).get('weight_kg', 0):.4f}kg",
                 f"{quote.get('material_cost', 0):.4f}", '含不良率'],
                ['加工费', f"{len(quote_data.get('process', {}).get('process_details', []))}工序",
                 f"{quote.get('process_cost', 0):.4f}", '含各工序'],
                ['管理费', f"{quote.get('rates', {}).get('management_rate', 0.045)*100:.1f}%",
                 f"{quote.get('management_cost', 0):.4f}", ''],
                ['其他费用', '包装+消耗品', f"{quote.get('other_cost', 0):.4f}", ''],
                ['利润', f"{quote.get('rates', {}).get('profit_rate', 0.15)*100:.1f}%",
                 f"{quote.get('profit', 0):.4f}", ''],
                ['单价总计', '', f"{quote.get('total_price', 0):.4f}", '元/件'],
            ]

            # 创建表格
            table = Table(table_data, colWidths=[40*mm, 50*mm, 40*mm, 50*mm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), font_name),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            # 绘制表格
            table.wrapOn(c, width, height)
            table.drawOn(c, 20*mm, y - 100*mm)

            y = y - 120*mm

            # 批量信息
            c.setFont(font_name, 12)
            c.drawString(30*mm, y, f"批量: {quote_data.get('lot_size', 2000)} 件")
            y -= 20
            c.setFont(font_name, 14)
            c.setFillColorRGB(1, 0, 0)
            c.drawString(30*mm, y, f"总金额: ¥{quote.get('total_price', 0) * quote_data.get('lot_size', 2000):,.2f}")
            c.setFillColorRGB(0, 0, 0)
            y -= 40

            # 备注
            c.setFont(font_name, 10)
            c.drawString(30*mm, y, "备注:")
            y -= 15
            c.drawString(35*mm, y, "1. 报价有效期30天")
            y -= 15
            c.drawString(35*mm, y, "2. 价格含税，不含运费")
            y -= 15
            c.drawString(35*mm, y, "3. 交货周期另议")

            # 页脚
            c.setFont(font_name, 8)
            c.drawCentredString(width/2, 20*mm, f"第 1 页 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

            c.save()
            output.seek(0)

            logger.info("PDF报价单生成成功")
            return output

        except Exception as e:
            logger.error(f"生成PDF报价单失败: {e}")
            raise ValueError(f"生成PDF失败: {e}")

    def generate_chenlong_template(
        self,
        customer_info: Dict,
        items: List[Dict],
        additional_info: Optional[Dict] = None
    ) -> BytesIO:
        """
        使用openpyxl生成报价单Excel（.xlsx格式）

        Args:
            customer_info: 客户信息
                - customer_name: 客户名称
                - contact_person: 联系人
                - phone: 电话
                - fax: 传真
                - quote_number: 报价单号
                - quote_date: 报价日期
                - company: 公司选择 ('shenzhen' 或 'dongguan'，默认 'shenzhen')
                - quoter: 报价人
            items: 产品列表
            additional_info: 附加信息（可选）

        Returns:
            Excel文件的BytesIO对象
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = '报价单'

            # ========== 定义样式 ==========
            # 颜色定义
            NAVY_BLUE = '1F4E79'      # 深蓝色 - 公司名称
            DARK_BLUE = '2E75B6'      # 蓝色 - 报价单标题
            HEADER_BG = '4472C4'      # 表头背景蓝色
            HEADER_BG_LIGHT = 'D6DCE5'  # 浅蓝灰色 - 子表头
            WHITE = 'FFFFFF'
            LIGHT_GRAY = 'F2F2F2'     # 斑马纹背景
            DARK_GRAY = '404040'      # 深灰色文字

            # 边框
            thin_border = Border(
                left=Side(style='thin', color='808080'),
                right=Side(style='thin', color='808080'),
                top=Side(style='thin', color='808080'),
                bottom=Side(style='thin', color='808080')
            )

            # 粗边框（用于盖章区域）
            medium_border = Border(
                left=Side(style='medium', color='404040'),
                right=Side(style='medium', color='404040'),
                top=Side(style='medium', color='404040'),
                bottom=Side(style='medium', color='404040')
            )

            # 字体 - 使用微软雅黑更现代美观
            title_font = Font(name='微软雅黑', size=20, bold=True, color=NAVY_BLUE)
            subtitle_font = Font(name='微软雅黑', size=16, bold=True, color=DARK_BLUE)
            info_label_font = Font(name='微软雅黑', size=10, bold=True, color=DARK_GRAY)
            info_value_font = Font(name='微软雅黑', size=10, color=DARK_GRAY)
            table_header_font = Font(name='微软雅黑', size=10, bold=True, color=WHITE)
            table_subheader_font = Font(name='微软雅黑', size=9, bold=True, color=DARK_GRAY)
            cell_font = Font(name='微软雅黑', size=10, color=DARK_GRAY)
            terms_font = Font(name='微软雅黑', size=9, color='595959')
            terms_title_font = Font(name='微软雅黑', size=10, bold=True, color=DARK_GRAY)
            sign_label_font = Font(name='微软雅黑', size=10, bold=True, color=DARK_GRAY)

            # 对齐
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')

            # 表头背景色
            header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
            subheader_fill = PatternFill(start_color=HEADER_BG_LIGHT, end_color=HEADER_BG_LIGHT, fill_type='solid')
            zebra_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

            # ========== 设置列宽 ==========
            col_widths = {'A': 6, 'B': 20, 'C': 10, 'D': 10, 'E': 14, 'F': 12, 'G': 10, 'H': 12, 'I': 12, 'J': 14}
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            # ========== 获取公司信息 ==========
            company_type = customer_info.get('company', 'shenzhen')
            if company_type == 'dongguan':
                company_name = '东莞市精之成精密五金有限公司'
                company_address = '地址：广东省东莞市横沥镇横沥育才路32号17号楼'
            else:
                company_name = '深圳市精之成精密五金有限公司'
                company_address = '地址：深圳市光明区马田街道石家社区下石家南环路东森工业区第三栋101、201'

            # ========== 第1行：公司名称 ==========
            ws.merge_cells('A1:J1')
            cell = ws['A1']
            cell.value = company_name
            cell.font = title_font
            cell.alignment = center_align
            ws.row_dimensions[1].height = 36

            # ========== 第2行：公司地址 ==========
            address_font = Font(name='微软雅黑', size=9, color='666666')
            ws.merge_cells('A2:J2')
            cell = ws['A2']
            cell.value = company_address
            cell.font = address_font
            cell.alignment = center_align
            ws.row_dimensions[2].height = 20

            # ========== 第3行：报价单标题 ==========
            ws.merge_cells('A3:J3')
            cell = ws['A3']
            cell.value = '报  价  单'
            cell.font = subtitle_font
            cell.alignment = center_align
            ws.row_dimensions[3].height = 28

            # ========== 第4行：分隔线 ==========
            ws.row_dimensions[4].height = 6

            # ========== 第5-7行：客户信息 ==========
            customer_name_val = customer_info.get('customer_name', '')
            contact_person = customer_info.get('contact_person', '')
            phone = customer_info.get('phone', '')
            fax = customer_info.get('fax', '')
            quote_number = customer_info.get('quote_number', datetime.now().strftime('%y%m%d'))
            date_str = customer_info.get('quote_date', datetime.now().strftime('%Y-%m-%d'))

            # 客户信息行
            ws.row_dimensions[5].height = 22
            ws.row_dimensions[6].height = 22
            ws.row_dimensions[7].height = 22

            ws.merge_cells('A5:D5')
            ws['A5'] = f'客户名称：{customer_name_val}'
            ws['A5'].font = info_value_font
            ws['A5'].alignment = left_align

            ws.merge_cells('G5:J5')
            ws['G5'] = f'日    期：{date_str}'
            ws['G5'].font = info_value_font
            ws['G5'].alignment = left_align

            ws.merge_cells('A6:D6')
            ws['A6'] = f'联 系 人：{contact_person}'
            ws['A6'].font = info_value_font
            ws['A6'].alignment = left_align

            ws.merge_cells('G6:J6')
            ws['G6'] = f'电    话：{phone}'
            ws['G6'].font = info_value_font
            ws['G6'].alignment = left_align

            ws.merge_cells('A7:D7')
            ws['A7'] = f'报价单号：{quote_number}'
            ws['A7'].font = info_label_font
            ws['A7'].alignment = left_align

            ws.merge_cells('G7:J7')
            ws['G7'] = f'传    真：{fax}'
            ws['G7'].font = info_value_font
            ws['G7'].alignment = left_align

            # ========== 第8行：空行 ==========
            ws.row_dimensions[8].height = 8

            # ========== 第9行：表头 ==========
            headers = ['序号', '客户料号', '直径', '长度', '材质', '表面处理', 'MOQ', '未税单价', '含税单价', '备注']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col_idx, value=header)
                cell.font = table_header_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

            ws.row_dimensions[9].height = 26

            # ========== 数据行（第10行开始） ==========
            start_row = 10
            total_rows = max(12, len(items))  # 至少12行，或者根据实际数据行数

            for idx, item in enumerate(items[:total_rows]):
                row_idx = start_row + idx
                ws.row_dimensions[row_idx].height = 22

                # 斑马纹背景（偶数行）
                row_fill = zebra_fill if idx % 2 == 1 else None

                # 序号 - 整数
                cell = ws.cell(row=row_idx, column=1, value=idx + 1)
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # 客户料号
                part_number = item.get('customer_part_number') or item.get('drawing_number', '')
                cell = ws.cell(row=row_idx, column=2, value=str(part_number) if part_number else '')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # 直径
                outer_diameter = item.get('outer_diameter', '')
                if outer_diameter:
                    try:
                        cell = ws.cell(row=row_idx, column=3, value=float(outer_diameter))
                        cell.number_format = '0.00'
                    except (ValueError, TypeError):
                        cell = ws.cell(row=row_idx, column=3, value=str(outer_diameter))
                else:
                    cell = ws.cell(row=row_idx, column=3, value='')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # 长度
                length = item.get('length', '')
                if length:
                    try:
                        cell = ws.cell(row=row_idx, column=4, value=float(length))
                        cell.number_format = '0.00'
                    except (ValueError, TypeError):
                        cell = ws.cell(row=row_idx, column=4, value=str(length))
                else:
                    cell = ws.cell(row=row_idx, column=4, value='')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # 材质
                material = item.get('material', '')
                cell = ws.cell(row=row_idx, column=5, value=str(material) if material else '')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # 表面处理
                surface_treatment = item.get('surface_treatment', '')
                cell = ws.cell(row=row_idx, column=6, value=str(surface_treatment) if surface_treatment else '')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # MOQ
                moq = item.get('lot_size') or item.get('quantity', '')
                if moq:
                    try:
                        cell = ws.cell(row=row_idx, column=7, value=int(float(moq)))
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        cell = ws.cell(row=row_idx, column=7, value=str(moq))
                else:
                    cell = ws.cell(row=row_idx, column=7, value='')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # 未税单价
                unit_price_before_tax = item.get('unit_price_before_tax', '')
                if unit_price_before_tax:
                    try:
                        cell = ws.cell(row=row_idx, column=8, value=float(unit_price_before_tax))
                        cell.number_format = '¥#,##0.0000'
                    except (ValueError, TypeError):
                        cell = ws.cell(row=row_idx, column=8, value=str(unit_price_before_tax))
                else:
                    cell = ws.cell(row=row_idx, column=8, value='')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # 含税单价
                unit_price_with_tax = item.get('unit_price_with_tax', '')
                if not unit_price_with_tax and unit_price_before_tax:
                    try:
                        unit_price_with_tax = round(float(unit_price_before_tax) * 1.13, 4)
                    except (ValueError, TypeError):
                        unit_price_with_tax = ''

                if unit_price_with_tax:
                    try:
                        cell = ws.cell(row=row_idx, column=9, value=float(unit_price_with_tax))
                        cell.number_format = '¥#,##0.0000'
                    except (ValueError, TypeError):
                        cell = ws.cell(row=row_idx, column=9, value=str(unit_price_with_tax))
                else:
                    cell = ws.cell(row=row_idx, column=9, value='')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                # 备注
                notes = item.get('notes', '')
                cell = ws.cell(row=row_idx, column=10, value=str(notes) if notes else '')
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

            # 填充空行（确保至少12行）
            for idx in range(len(items), total_rows):
                row_idx = start_row + idx
                ws.row_dimensions[row_idx].height = 22
                row_fill = zebra_fill if idx % 2 == 1 else None

                cell = ws.cell(row=row_idx, column=1, value=idx + 1)
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                for col_idx in range(2, 11):
                    cell = ws.cell(row=row_idx, column=col_idx, value='')
                    cell.font = cell_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill

            # ========== 附加条款区域 ==========
            footer_row = start_row + total_rows + 1
            ws.row_dimensions[footer_row].height = 12  # 空行

            # 条款标题
            footer_row += 1
            ws.merge_cells(f'A{footer_row}:J{footer_row}')
            cell = ws.cell(row=footer_row, column=1, value='【报价条款】')
            cell.font = terms_title_font
            cell.alignment = left_align
            ws.row_dimensions[footer_row].height = 20

            footer_row += 1
            ws.row_dimensions[footer_row].height = 18
            # 左侧条款
            ws.merge_cells(f'A{footer_row}:E{footer_row}')
            cell = ws.cell(row=footer_row, column=1, value='  1) 报价有效期: 30天')
            cell.font = terms_font
            cell.alignment = left_align

            # 右侧条款
            ws.merge_cells(f'F{footer_row}:J{footer_row}')
            cell = ws.cell(row=footer_row, column=6, value='  4) 生产周期: 30天~2个月 F/C')
            cell.font = terms_font
            cell.alignment = left_align

            footer_row += 1
            ws.row_dimensions[footer_row].height = 18
            ws.merge_cells(f'A{footer_row}:E{footer_row}')
            cell = ws.cell(row=footer_row, column=1, value='  2) 交货地点: 贵公司仓库')
            cell.font = terms_font
            cell.alignment = left_align

            ws.merge_cells(f'F{footer_row}:J{footer_row}')
            cell = ws.cell(row=footer_row, column=6, value='  5) 以上报价含税13%')
            cell.font = terms_font
            cell.alignment = left_align

            footer_row += 1
            ws.row_dimensions[footer_row].height = 18
            ws.merge_cells(f'A{footer_row}:E{footer_row}')
            cell = ws.cell(row=footer_row, column=1, value='  3) 付款条件: 月结60天')
            cell.font = terms_font
            cell.alignment = left_align

            ws.merge_cells(f'F{footer_row}:J{footer_row}')
            cell = ws.cell(row=footer_row, column=6, value='  6) 本报价依据为贵公司提供之图纸')
            cell.font = terms_font
            cell.alignment = left_align

            # ========== 签章区域 ==========
            footer_row += 1
            ws.row_dimensions[footer_row].height = 20  # 空行

            footer_row += 1
            quoter = customer_info.get('quoter', '')

            # 定义签章区域的边框样式
            sign_border = Border(
                left=Side(style='thin', color='808080'),
                right=Side(style='thin', color='808080'),
                top=Side(style='thin', color='808080'),
                bottom=Side(style='thin', color='808080')
            )
            sign_header_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

            # ===== 报价人区域 (A-C列) =====
            # 标题行
            ws.merge_cells(f'A{footer_row}:C{footer_row}')
            cell = ws.cell(row=footer_row, column=1, value='报 价 人')
            cell.font = Font(name='微软雅黑', size=10, bold=True, color=DARK_GRAY)
            cell.alignment = center_align
            cell.fill = sign_header_fill
            for c in range(1, 4):
                ws.cell(row=footer_row, column=c).border = sign_border

            # ===== 供方盖章区域 (D-G列) =====
            ws.merge_cells(f'D{footer_row}:G{footer_row}')
            cell = ws.cell(row=footer_row, column=4, value='供 方 盖 章')
            cell.font = Font(name='微软雅黑', size=10, bold=True, color=DARK_GRAY)
            cell.alignment = center_align
            cell.fill = sign_header_fill
            for c in range(4, 8):
                ws.cell(row=footer_row, column=c).border = sign_border

            # ===== 客户确认区域 (H-J列) =====
            ws.merge_cells(f'H{footer_row}:J{footer_row}')
            cell = ws.cell(row=footer_row, column=8, value='客 户 确 认')
            cell.font = Font(name='微软雅黑', size=10, bold=True, color=DARK_GRAY)
            cell.alignment = center_align
            cell.fill = sign_header_fill
            for c in range(8, 11):
                ws.cell(row=footer_row, column=c).border = sign_border

            ws.row_dimensions[footer_row].height = 22

            # ===== 签章内容区域 =====
            footer_row += 1
            sign_content_rows = 4  # 签章区域高度

            # 报价人签名框
            ws.merge_cells(f'A{footer_row}:C{footer_row + sign_content_rows - 1}')
            cell = ws.cell(row=footer_row, column=1, value=quoter if quoter else '')
            cell.font = Font(name='微软雅黑', size=12, color=DARK_GRAY)
            cell.alignment = center_align
            for r in range(footer_row, footer_row + sign_content_rows):
                for c in range(1, 4):
                    ws.cell(row=r, column=c).border = sign_border

            # 公司盖章框
            ws.merge_cells(f'D{footer_row}:G{footer_row + sign_content_rows - 1}')
            cell = ws.cell(row=footer_row, column=4, value='')
            cell.alignment = center_align
            for r in range(footer_row, footer_row + sign_content_rows):
                for c in range(4, 8):
                    ws.cell(row=r, column=c).border = sign_border

            # 客户盖章框
            ws.merge_cells(f'H{footer_row}:J{footer_row + sign_content_rows - 1}')
            cell = ws.cell(row=footer_row, column=8, value='')
            cell.alignment = center_align
            for r in range(footer_row, footer_row + sign_content_rows):
                for c in range(8, 11):
                    ws.cell(row=r, column=c).border = sign_border

            # 设置签章区域行高
            for r in range(footer_row, footer_row + sign_content_rows):
                ws.row_dimensions[r].height = 22

            # ===== 日期行 =====
            footer_row += sign_content_rows
            ws.merge_cells(f'A{footer_row}:C{footer_row}')
            cell = ws.cell(row=footer_row, column=1, value='日期：')
            cell.font = Font(name='微软雅黑', size=9, color='808080')
            cell.alignment = center_align
            for c in range(1, 4):
                ws.cell(row=footer_row, column=c).border = sign_border

            ws.merge_cells(f'D{footer_row}:G{footer_row}')
            cell = ws.cell(row=footer_row, column=4, value='日期：')
            cell.font = Font(name='微软雅黑', size=9, color='808080')
            cell.alignment = center_align
            for c in range(4, 8):
                ws.cell(row=footer_row, column=c).border = sign_border

            ws.merge_cells(f'H{footer_row}:J{footer_row}')
            cell = ws.cell(row=footer_row, column=8, value='日期：')
            cell.font = Font(name='微软雅黑', size=9, color='808080')
            cell.alignment = center_align
            for c in range(8, 11):
                ws.cell(row=footer_row, column=c).border = sign_border

            ws.row_dimensions[footer_row].height = 20

            # 保存到BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info(f"✅ 报价单生成成功，包含{len(items)}个产品，公司：{company_name}")
            return output

        except Exception as e:
            logger.error(f"❌ 生成报价单失败: {e}")
            import traceback
            traceback.print_exc()
            raise ValueError(f"生成Excel失败: {e}")


# 全局单例
_generator = None


def get_document_generator() -> QuoteDocumentGenerator:
    """获取文档生成器单例"""
    global _generator
    if _generator is None:
        _generator = QuoteDocumentGenerator()
    return _generator
